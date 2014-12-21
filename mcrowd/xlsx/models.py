import json
import logging
import os
import pickle
import stat
import io

from django.contrib.auth.models import User
from django.db import models
from django.db.models.signals import post_save, pre_save

from .hooks import check_header, create_worksheets
from .hooks import load_native_workbook, save_to_cache
from .settings import LOCALSTORE, HEADER_LOCATION, DATA_LOCATION
from .utils import get_data_rows

from mcrowd.common.exceptions import BadRequest

import openpyxl

logger = logging.getLogger(__name__)


class Workbook(models.Model):
    owner = models.ForeignKey(User, null=False)
    data = models.BinaryField(null=False, blank=False)
    filename = models.TextField(null=False, blank=False)

    def get_workbook(self):
        filename = self.get_cached_name(self.pk)
        if not filename:
            book = self.load_workbook()
            self.save_to_cache()
        else:
            book = openpyxl.load_workbook(filename)
        return book

    @staticmethod
    def get_cached_name(pk):
        return os.path.join(LOCALSTORE, "{}.xlsx".format(pk))

    def load_workbook(self):
        data = io.BytesIO(bytes(self.data))
        try:
            return openpyxl.load_workbook(data)
        except Exception as error:
            logger.exception(error)
            raise BadRequest(
                detail="Could not load xlsx with name {}".format(obj.name))

    def save_to_cache(self):
        with open(self.get_cached_name(self.pk), 'wb') as h:
            h.write(bytes(self.data))

    def __unicode__(self):
        return self.filename

    class Meta:
        unique_together = (("owner", "filename"),)


class Worksheet(models.Model):
    workbook = models.ForeignKey(Workbook, null=False, related_name="sheets")
    name = models.TextField(blank=False, null=False, db_index=True)
    number = models.IntegerField(blank=False, null=False)

    def get_worksheet(self):
        filename = self.get_cached_name(self.pk)
        if not os.path.exists(filename):
            book = self.workbook.get_workbook()
            index = book.get_sheet_names().index(self.name)
            sheet = book.worksheets[index]
            self.save_to_cache(sheet)
            return sheet
        with open(filename, 'rb') as h:
            return pickle.load(h)

    @staticmethod
    def get_cached_name(pk):
        return os.path.join(LOCALSTORE, "sheet-{}.pkl".format(pk))

    def save_to_cache(self, sheet):
        with open(self.get_cached_name(self.pk), 'wb') as h:
            pickle.dump(sheet, h, pickle.HIGHEST_PROTOCOL)

    def __unicode__(self):
        return "{}!{}".format(self.workbook.__unicode__(), self.name)


class Table(models.Model):
    worksheet = models.ForeignKey(Worksheet, null=False)
    col_names = models.TextField(blank=False, null=False, default="[]")
    col_ids = models.TextField(blank=False, null=False, default="[]")
    header_location = models.CharField(
        blank=False, null=False, max_length=20, default=HEADER_LOCATION)
    data_location = models.CharField(
        blank=False, null=False, max_length=20, default=DATA_LOCATION)

    def get_col_names(self):
        return json.loads(self.col_names)

    def get_col_ids(self):
        return json.loads(self.col_ids)

    def get_index(self, name):
        index = self.get_col_names().index(name)
        return self.get_col_ids()[index]

    def get_rows(self, col_ids=None, limit=None):
        sheet = self.worksheet.get_worksheet()
        col_ids = col_ids or self.get_col_ids()
        rows = []
        for i, (number, row) in enumerate(
                get_data_rows(sheet, col_ids, self.data_location)):
            row = sorted(row, key=lambda x: self.get_col_ids().index(x.column))
            row = list(map(lambda x: x.value or "", row))
            rows.append([number, row])
            if limit is not None and i >= limit:
                break
        return rows

    def __unicode__(self):
        return "{}: {}".format(self.worksheet.__unicode__(), self.col_names)


pre_save.connect(load_native_workbook, Workbook)
post_save.connect(save_to_cache, Workbook)
post_save.connect(create_worksheets, Workbook)
pre_save.connect(check_header, Table)
