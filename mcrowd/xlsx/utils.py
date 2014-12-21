import logging

from openpyxl.cell import coordinate_from_string
from openpyxl.cell import column_index_from_string, get_column_letter

from mcrowd.common.exceptions import BadRequest

logger = logging.getLogger(__name__)


def get_square_indices(sheet, location):
    left, right = location.upper().split(':')
    left_column_letter, left_row = coordinate_from_string(left)
    left_column = column_index_from_string(left_column_letter)
    if right:
        right_column_letter, right_row = coordinate_from_string(right)
        right_column = column_index_from_string(right_column_letter)
        right_defined = True
    else:
        right_row = sheet.get_highest_row()
        right_column = sheet.get_highest_column()
        right_defined = False
    if left_column > right_column:
        raise BadRequest(detail="Left column should be less than right")
    if left_row <= 0 or right_row <= 0:
        raise BadRequest(detail="Row index should be greater than 0")
    if left_row > right_row:
        raise BadRequest(
            detail="Left row index should be less or equal than right row")
    return left_column, left_row, right_column, right_row, right_defined


def get_range(sheet, left_column, left_row, right_column, right_row):
    range_string = "%s%d:%s%d" % (get_column_letter(left_column), left_row,
                                  get_column_letter(right_column), right_row)
    return sheet.range(range_string)


def get_header_columns(sheet, header_location):
    left_column, left_row, right_column, right_row, right_defined = \
        get_square_indices(sheet, header_location)
    if not right_defined:
        right_row = left_row
    if right_row != left_row:
        raise BadRequest(detail="Multiline headers are not implemented yet")
    header = get_range(sheet, left_column, left_row, right_column, right_row)
    for row in header:
        return row


def get_header_index_by_name(header, name):
    found = list(filter(lambda x: str(x.value) == str(name), header))
    if found:
        return found[0].column
    raise BadRequest(
        detail="Could not find column with name: '%s'" % name)


def is_empty_row(row):
    not_empty = list(filter(lambda x: x.value, row))
    return not bool(not_empty)


def get_data_rows(sheet, col_ids, data_location):
    left_column, left_row, right_column, right_row, right_defined = \
        get_square_indices(sheet, data_location)
    left_column = max(left_column, column_index_from_string(min(*col_ids)))
    right_column = min(right_column, column_index_from_string(max(*col_ids)))
    data = get_range(sheet, left_column, left_row, right_column, right_row)
    for i, row in enumerate(data):
        selected = tuple(filter(lambda x: x.column in col_ids, row))
        yield left_row + i, selected
